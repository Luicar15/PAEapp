import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import win32com.client as win32
from datetime import datetime
from PyPDF2 import PdfMerger
import pythoncom

PLANTILLA_REMISION = os.path.join("plantillas", "remision.xlsx")
ARCHIVO_BASE = os.path.join("datos_base_completo_final_v2.xlsx")

# ---------- FUNCIONES AUXILIARES ---------- #

def cargar_insumos_por_menu_y_grupo(hoja_complemento, menu_num, grupo_etario, raciones):
    """Carga alimentos desde el archivo base para un menú y grupo etario."""
    df = pd.read_excel(ARCHIVO_BASE, sheet_name=hoja_complemento)
    seleccion = df.loc[(df["menu"] == menu_num) & (df["grupo_etario"] == grupo_etario)].copy()
    if seleccion.empty:
        return pd.DataFrame(columns=["alimento", "unidad", "grupo_etario", "cantidad_total"])
    seleccion["grupo_etario"] = grupo_etario
    seleccion["cantidad_total"] = (seleccion["cantidad_por_racion"] * raciones).round(2)
    return seleccion[["alimento", "unidad", "grupo_etario", "cantidad_total"]]

def consolidar_insumos_remision(hoja_complemento, menus_dias, grupos_raciones):
    """Genera insumos acumulados por alimento, grupo etario y día."""
    registros = []
    for dia, menu_num in menus_dias.items():
        for grupo, raciones in grupos_raciones.items():
            if pd.notna(menu_num) and menu_num != 0 and raciones > 0:
                insumos = cargar_insumos_por_menu_y_grupo(
                    hoja_complemento, menu_num, grupo, raciones
                )
                if not insumos.empty:
                    insumos["dia"] = dia
                    registros.append(insumos)
    if not registros:
        return pd.DataFrame(
            columns=["alimento", "unidad", "grupo_etario", "cantidad_total", "dia"]
        )
    return pd.concat(registros, ignore_index=True)


def pivotar_por_grupo(insumos: pd.DataFrame, grupos: list[str]) -> pd.DataFrame:
    """Transforma los insumos para tener una fila por alimento con columnas por grupo."""
    if insumos.empty:
        cols = ["alimento", "unidad", "cantidad_total"] + grupos
        return pd.DataFrame(columns=cols)

    tabla = (
        insumos.pivot_table(
            index=["alimento", "unidad"],
            columns="grupo_etario",
            values="cantidad_total",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    for g in grupos:
        if g not in tabla.columns:
            tabla[g] = 0

    tabla["cantidad_total"] = tabla[grupos].sum(axis=1)
    return tabla

def aplicar_formato_fila(hoja, fila, alto=20.25):
    font = Font(name="Arial", size=8, bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    hoja.row_dimensions[fila].height = alto
    for col in range(1, 9):
        celda = hoja.cell(row=fila, column=col)
        celda.font = font
        celda.alignment = align
        celda.border = border

def aplicar_bloque_observaciones_firmas(hoja, fila_inicio):
    textos = [
        ("OBSERVACIONES:", ""),
        ("NOMBRE DEL RESPONSABLE DEL OPERADOR:", "FIRMA:"),
        ("NOMBRE DEL RESPONSABLE DE LA INSTITUCIÓN:", "FIRMA:"),
        ("CARGO:", "FIRMA:")
    ]
    font = Font(name="Arial", size=8, bold=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for i, (texto_a, texto_h) in enumerate(textos):
        fila = fila_inicio + i
        hoja.merge_cells(f"A{fila}:D{fila}")
        hoja.merge_cells(f"E{fila}:H{fila}")
        hoja[f"A{fila}"] = texto_a
        hoja[f"E{fila}"] = texto_h
        hoja.row_dimensions[fila].height = 24.5
        for col in range(1, 9):
            celda = hoja.cell(row=fila, column=col)
            celda.font = font
            celda.alignment = align_left
            celda.border = border

def limpiar_celda_si_combinada(hoja, ref):
    for rango in list(hoja.merged_cells.ranges):
        if ref in rango:
            hoja.unmerge_cells(str(rango))
            break

def convertir_excels_a_pdfs(lista_excels):
    pythoncom.CoInitialize()
    excel_app = win32.DispatchEx("Excel.Application")
    pdfs = []
    for xlsx in lista_excels:
        pdf = xlsx.replace(".xlsx", ".pdf")
        wb = excel_app.Workbooks.Open(os.path.abspath(xlsx))
        ws = wb.ActiveSheet
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False
        ws.PageSetup.TopMargin = 1.91 / 2.54
        ws.PageSetup.BottomMargin = 1.91 / 2.54
        ws.PageSetup.LeftMargin = 1.78 / 2.54
        ws.PageSetup.RightMargin = 1.78 / 2.54
        ws.PageSetup.HeaderMargin = 0.76 / 2.54
        ws.PageSetup.FooterMargin = 0.76 / 2.54
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf))
        wb.Close(False)
        pdfs.append(pdf)
    excel_app.Quit()
    return pdfs

def unir_pdfs(lista_pdfs, ruta_pdf_final):
    merger = PdfMerger()
    for pdf in lista_pdfs:
        merger.append(pdf)
    merger.write(ruta_pdf_final)
    merger.close()

# ---------- GENERACIÓN DE REMISIÓN ---------- #

def generar_excel_institucion_remision(fila, carpeta_salida):
    wb = load_workbook(PLANTILLA_REMISION)
    hoja = wb.active

    dias_atendidos = fila.get("Días Atendidos", fila.get("Dias Atendidos", 1))

    hoja["K2"] = pd.to_datetime(fila["Fecha Inicial"]).strftime("%d/%m/%Y") if not pd.isna(fila["Fecha Inicial"]) else ""
    hoja["J3"] = fila["Municipio"]
    hoja["B4"] = fila["Institucion Educativa (IED)"]
    hoja["K4"] = fila["Sede"]

    # Raciones por grupo
    grados = [
        ("Grado 0 (Raciones)", "D7", "G7", "J7"),
        ("Grado 1,2,3 (Raciones)", "D8", "G8", "J8"),
        ("Grado 4,5 (Raciones)", "D9", "G9", "J9"),
        ("Grado 6,7,8,9 (Raciones)", "D10", "G10", "J10"),
        ("Grado 10,11 (Raciones)", "D11", "G11", "J11"),
    ]
    total_raciones = 0
    for grado, cel_r, cel_calc, cel_dias in grados:
        valor = fila[grado] if pd.notna(fila[grado]) else 0
        hoja[cel_r] = valor
        hoja[cel_calc] = valor * dias_atendidos
        hoja[cel_dias] = dias_atendidos
        total_raciones += valor * dias_atendidos

    hoja["M7"] = " | ".join(str(fila.get(k, "")) for k in [
        "Lunes (Menú)", "Martes (Menú)", "Miercoles (Menú)",
        "Jueves (Menú)", "Viernes (Menú)", "Ciclo de Menú (Semana)"
    ])
    hoja["P7"] = f"{fila['Tipo de Complemento']} - Total Raciones: {total_raciones}"

    # Consolidar insumos reales por menú y grupo etario
    grupos_raciones = {g: fila[g] for g, _, _, _ in grados}
    menus_dias = {
        "lunes": fila["Lunes (Menú)"],
        "martes": fila["Martes (Menú)"],
        "miercoles": fila["Miercoles (Menú)"],
        "jueves": fila["Jueves (Menú)"],
        "viernes": fila["Viernes (Menú)"]
    }
    hoja_complemento = fila["Complemento Alimenticio"]
    insumos = consolidar_insumos_remision(
        hoja_complemento, menus_dias, grupos_raciones
    )

    grupos_columnas = list(grupos_raciones.keys())
    resumen = pivotar_por_grupo(insumos, grupos_columnas)

    # Insertar en Excel desde la fila 15
    fila_inicio = 15
    for _, row in resumen.iterrows():
        limpiar_celda_si_combinada(hoja, f"A{fila_inicio}")
        limpiar_celda_si_combinada(hoja, f"G{fila_inicio}")
        limpiar_celda_si_combinada(hoja, f"H{fila_inicio}")

        hoja[f"A{fila_inicio}"] = row["alimento"]
        hoja[f"G{fila_inicio}"] = row["unidad"]
        hoja[f"H{fila_inicio}"] = round(row["cantidad_total"], 2)

        for idx, grupo in enumerate(grupos_columnas):
            valor = round(row.get(grupo, 0), 2)
            hoja.cell(row=fila_inicio, column=2 + idx).value = valor

        aplicar_formato_fila(hoja, fila_inicio)
        fila_inicio += 1

    aplicar_bloque_observaciones_firmas(hoja, fila_inicio + 1)

    nombre = f"Remision_{fila['Institucion Educativa (IED)'].replace(' ', '_')}_{fila['Sede'].replace(' ', '_')}.xlsx"
    ruta_excel = os.path.join(carpeta_salida, nombre)
    if os.path.exists(ruta_excel):
        os.remove(ruta_excel)
    wb.save(ruta_excel)
    return ruta_excel

def generar_remision_consolidado(df_instituciones, progreso, carpeta_salida):
    excels = []
    for _, fila in df_instituciones.iterrows():
        excels.append(generar_excel_institucion_remision(fila, carpeta_salida))
        progreso["procesadas"] += 1
        time.sleep(1)

    pdfs = convertir_excels_a_pdfs(excels)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    ruta_pdf_final = os.path.join(carpeta_salida, f"Remision_PAE_{fecha}.pdf")
    unir_pdfs(pdfs, ruta_pdf_final)

    for p in pdfs:
        if os.path.exists(p):
            os.remove(p)

    return ruta_pdf_final, excels