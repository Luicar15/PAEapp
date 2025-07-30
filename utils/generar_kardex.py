import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import win32com.client as win32
from datetime import datetime
from PyPDF2 import PdfMerger
import pythoncom

# Archivos base
PLANTILLA_KARDEX = os.path.join("plantillas", "kardex.xlsx")
ARCHIVO_BASE = os.path.join("datos_base_completo_final_v2.xlsx")

# ---------------- FUNCIONES AUXILIARES ---------------- #

def cargar_insumos(hoja_complemento, menu_num, grupo_etario, raciones):
    df = pd.read_excel(ARCHIVO_BASE, sheet_name=hoja_complemento)
    seleccion = df.loc[(df["menu"] == menu_num) & (df["grupo_etario"] == grupo_etario)].copy()
    if seleccion.empty:
        return pd.DataFrame(columns=["alimento", "unidad", "cantidad_total"])
    seleccion["cantidad_total"] = (seleccion["cantidad_por_racion"] * raciones).round(2)
    return seleccion[["alimento", "unidad", "cantidad_total"]]

def consolidar_insumos(hoja_complemento, menus_dias, grupos_raciones):
    totales = pd.DataFrame()
    diarios = {dia: pd.DataFrame() for dia in menus_dias}
    for dia, menu_num in menus_dias.items():
        insumos_dia = []
        for grupo, raciones in grupos_raciones.items():
            if pd.notna(menu_num) and menu_num != 0:
                insumos = cargar_insumos(hoja_complemento, menu_num, grupo, raciones)
                if not insumos.empty:
                    insumos_dia.append(insumos)
        if insumos_dia:
            diarios[dia] = pd.concat(insumos_dia).groupby(
                ["alimento", "unidad"], as_index=False).sum(numeric_only=True)
            diarios[dia]["cantidad_total"] = diarios[dia]["cantidad_total"].round(2)
    if diarios:
        totales = pd.concat(list(diarios.values())).groupby(
            ["alimento", "unidad"], as_index=False).sum(numeric_only=True)
        totales["cantidad_total"] = totales["cantidad_total"].round(2)
    return totales, diarios

def aplicar_formato_fila(hoja, fila, alto=20.25):
    font = Font(name="Arial", size=8, bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    hoja.row_dimensions[fila].height = alto
    for col in range(1, 16):  # Columnas A (1) a O (15)
        celda = hoja.cell(row=fila, column=col)
        celda.font = font
        celda.alignment = align
        celda.border = border

def aplicar_formato_firmas(hoja, fila_inicio):
    """Crea y formatea el bloque de firmas a partir de la fila dada (alineado a la izquierda)."""
    textos = [
        ("NOMBRE DEL TRANSPORTADOR (Operador):", "FIRMA:"),
        ("NOMBRE MANIPULADOR DE ALIMENTOS QUE RECIBE (Operador):",
         "NOMBRE RESPONSABLE INSTITUCIÓN O CENTRO EDUCATIVO:"),
        ("FIRMA:", "CARGO:"),
        ("FIRMA:", "FIRMA:")
    ]

    font = Font(name="Arial", size=8, bold=True)
    align_izquierda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for i, (texto_a, texto_h) in enumerate(textos):
        fila = fila_inicio + i

        # Combinar correctamente columnas
        hoja.merge_cells(f"A{fila}:G{fila}")
        hoja.merge_cells(f"H{fila}:O{fila}")

        celda_a = hoja[f"A{fila}"]
        celda_h = hoja[f"H{fila}"]

        # Asignar textos
        celda_a.value = texto_a
        celda_h.value = texto_h

        # Aplicar estilo a toda la fila A:O
        for col in range(1, 16):
            celda = hoja.cell(row=fila, column=col)
            celda.font = font
            celda.alignment = align_izquierda
            celda.border = border

        hoja.row_dimensions[fila].height = 19.5

def limpiar_celda_si_combinada(hoja, ref):
    for rango in list(hoja.merged_cells.ranges):
        if ref in rango:
            hoja.unmerge_cells(str(rango))
            break

def convertir_excels_a_pdfs(lista_excels):
    pythoncom.CoInitialize()
    excel_app = win32.DispatchEx("Excel.Application")
    pdfs = []
    for xlsx in lista_excels:
        pdf = xlsx.replace(".xlsx", ".pdf")
        wb = excel_app.Workbooks.Open(os.path.abspath(xlsx))
        ws = wb.ActiveSheet

        # Configurar centrado horizontal y márgenes (cm -> pulgadas)
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False
        ws.PageSetup.TopMargin = 1.91 / 2.54
        ws.PageSetup.BottomMargin = 1.91 / 2.54
        ws.PageSetup.LeftMargin = 1.78 / 2.54
        ws.PageSetup.RightMargin = 1.78 / 2.54
        ws.PageSetup.HeaderMargin = 0.76 / 2.54
        ws.PageSetup.FooterMargin = 0.76 / 2.54

        wb.ExportAsFixedFormat(0, os.path.abspath(pdf))
        wb.Close(False)
        pdfs.append(pdf)
    excel_app.Quit()
    return pdfs

def unir_pdfs(lista_pdfs, ruta_pdf_final):
    merger = PdfMerger()
    for pdf in lista_pdfs:
        merger.append(pdf)
    merger.write(ruta_pdf_final)
    merger.close()

# ---------------- GENERACIÓN DE KARDEX ---------------- #

def generar_excel_institucion(fila, carpeta_salida):
    wb = load_workbook(PLANTILLA_KARDEX)
    hoja = wb.active

    # Insertar datos fijos
    hoja["I3"] = fila["Municipio"]
    hoja["C4"] = fila["Institucion Educativa (IED)"]
    hoja["J4"] = fila["Sede"]
    hoja["C5"] = fila["Complemento Alimenticio"]
    hoja["G5"] = fila["Tipo de Complemento"]
    hoja["C6"] = fila["Grado 0 (Raciones)"]
    hoja["F6"] = fila["Grado 1,2,3 (Raciones)"]
    hoja["I6"] = fila["Grado 4,5 (Raciones)"]
    hoja["L6"] = fila["Grado 6,7,8,9 (Raciones)"]
    hoja["O6"] = fila["Grado 10,11 (Raciones)"]
    hoja["F7"] = fila["Cupos Totales"]
    hoja["K2"] = pd.to_datetime(fila["Fecha Inicial"]).strftime("%d/%m/%Y") if not pd.isna(fila["Fecha Inicial"]) else ""
    hoja["J7"] = fila["Ciclo de Menú (Semana)"]

    # Configurar raciones y menús
    grupos_raciones = {
        "Grado 0 (Raciones)": fila["Grado 0 (Raciones)"],
        "Grado 1,2,3 (Raciones)": fila["Grado 1,2,3 (Raciones)"],
        "Grado 4,5 (Raciones)": fila["Grado 4,5 (Raciones)"],
        "Grado 6,7,8,9 (Raciones)": fila["Grado 6,7,8,9 (Raciones)"],
        "Grado 10,11 (Raciones)": fila["Grado 10,11 (Raciones)"]
    }
    menus_dias = {
        "lunes": fila["Lunes (Menú)"],
        "martes": fila["Martes (Menú)"],
        "miercoles": fila["Miercoles (Menú)"],
        "jueves": fila["Jueves (Menú)"],
        "viernes": fila["Viernes (Menú)"]
    }

    hoja_complemento = fila["Complemento Alimenticio"]
    totales, diarios = consolidar_insumos(hoja_complemento, menus_dias, grupos_raciones)

    # Insertar alimentos
    fila_inicio = 9
    for _, row in totales.iterrows():
        for col_ref, valor in zip(["B", "C", "D"], [row["alimento"], round(row["cantidad_total"], 2), row["unidad"]]):
            ref = f"{col_ref}{fila_inicio}"
            limpiar_celda_si_combinada(hoja, ref)
            hoja[ref] = valor
        for idx, dia in enumerate(["lunes", "martes", "miercoles", "jueves", "viernes"]):
            if dia in diarios and not diarios[dia].empty:
                match = diarios[dia][diarios[dia]["alimento"] == row["alimento"]]
                valor = round(match["cantidad_total"].values[0], 2) if not match.empty else 0
                col = ["E", "G", "I", "K", "M"][idx]
                ref = f"{col}{fila_inicio}"
                limpiar_celda_si_combinada(hoja, ref)
                hoja[ref] = valor
        aplicar_formato_fila(hoja, fila_inicio)
        fila_inicio += 1

    # Insertar bloque de firmas (comienza en la fila siguiente al último alimento)
    aplicar_formato_firmas(hoja, fila_inicio + 1)

    # Guardar Excel por institución
    nombre = f"Kardex_{fila['Institucion Educativa (IED)'].replace(' ', '_')}_{fila['Sede'].replace(' ', '_')}.xlsx"
    ruta_excel = os.path.join(carpeta_salida, nombre)
    if os.path.exists(ruta_excel):
        try:
            os.remove(ruta_excel)
        except PermissionError:
            os.system(f"taskkill /f /im EXCEL.EXE")
            time.sleep(1)
            if os.path.exists(ruta_excel):
                os.remove(ruta_excel)
    wb.save(ruta_excel)
    return ruta_excel

def generar_kardex_consolidado(df_instituciones, progreso, carpeta_salida):
    excels = []
    for _, fila in df_instituciones.iterrows():
        excels.append(generar_excel_institucion(fila, carpeta_salida))
        progreso["procesadas"] += 1
        time.sleep(1)

    pdfs = convertir_excels_a_pdfs(excels)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    ruta_pdf_final = os.path.join(carpeta_salida, f"Kardex_PAE_{fecha}.pdf")
    unir_pdfs(pdfs, ruta_pdf_final)

    for p in pdfs:
        if os.path.exists(p):
            os.remove(p)

    return ruta_pdf_final, excels