import os
import pandas as pd
import openpyxl
import win32com.client as win32
from tkinter import Tk, filedialog
from database import get_connection

# Rutas de las plantillas oficiales
PLANTILLAS = {
    "kardex": "plantillas/1. FORMATO KARDEX.xlsx",
    "remision": "plantillas/2. FORMATO REMISIÓN.xlsx",
    "certificacion": "plantillas/4. 358_formato--certificado-de-entrega-de-raciones-en-instituciones-educativas.xlsx",
    "asistencia": "plantillas/5. 364_formato--registro-y-control-diario-de-asistencia (1).xls"
}

# Fuente de datos base para insumos y ciclos
DATOS_BASE = "datos_base_completo_final_v2.xlsx"


def seleccionar_directorio_guardado():
    """Abre un diálogo para que el usuario seleccione dónde guardar los PDFs."""
    Tk().withdraw()
    carpeta = filedialog.askdirectory(title="Seleccione carpeta para guardar los documentos generados")
    return carpeta


def calcular_insumos(complemento, ciclo, grupo, raciones):
    """Calcula insumos según el menú usando datos_base_completo_final_v2.xlsx"""
    df_menus = pd.read_excel(DATOS_BASE, sheet_name="menus")
    filtro = (df_menus['componente'] == complemento) & (df_menus['menu'] == ciclo) & (df_menus['grupo'] == grupo)
    seleccion = df_menus.loc[filtro].copy()

    if seleccion.empty:
        return []

    seleccion['cantidad_total'] = seleccion['cantidad_por_racion'] * raciones
    return seleccion[['categoria', 'alimento', 'codigo', 'cantidad_por_racion', 'unidad', 'cantidad_total']].to_dict('records')


def generar_documentos(df):
    """Genera 4 archivos PDF (Kardex, Remisión, Certificación, Registro) para todas las instituciones del cargue."""
    carpeta_salida = seleccionar_directorio_guardado()
    if not carpeta_salida:
        return "Operación cancelada."

    # Crear una copia de cada plantilla y llenarla con datos de todas las filas
    archivos_salida = {}
    for tipo, plantilla in PLANTILLAS.items():
        salida_xlsx = os.path.join(carpeta_salida, f"{tipo.upper()}_PAE.xlsx")
        salida_pdf = os.path.join(carpeta_salida, f"{tipo.upper()}_PAE.pdf")
        archivos_salida[tipo] = salida_pdf

        wb = openpyxl.load_workbook(plantilla)
        ws = wb.active

        # Insertar datos (ejemplo: solo encabezados y totales, la lógica completa se adapta al diseño real)
        fila = 10
        for _, row in df.iterrows():
            ws.cell(row=fila, column=1).value = row['Municipio']
            ws.cell(row=fila, column=2).value = row['Institucion Educativa (IED)']
            ws.cell(row=fila, column=3).value = row['Sede']
            ws.cell(row=fila, column=4).value = row['Complemento Alimenticio']
            ws.cell(row=fila, column=5).value = row['Cupos Totales']
            ws.cell(row=fila, column=6).value = row['Días Atendidos']
            fila += 1

        wb.save(salida_xlsx)

        # Convertir a PDF usando Excel
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        wb_excel = excel.Workbooks.Open(os.path.abspath(salida_xlsx))
        wb_excel.ExportAsFixedFormat(0, os.path.abspath(salida_pdf))
        wb_excel.Close()
        excel.Quit()

    return f"Documentos generados correctamente en {carpeta_salida}"


def procesar_cargue_masivo(archivo_excel):
    """Procesa el Excel maestro, calcula insumos y genera PDFs consolidados."""
    df = pd.read_excel(archivo_excel)

    # Calcular insumos para cada fila
    resultados = []
    for _, row in df.iterrows():
        insumos = []
        for grupo, col_raciones in zip(
            ['PREESCOLAR', 'PRIMARIA_BAJA', 'PRIMARIA_ALTA', 'SECUNDARIA', 'MYC'],
            ['Grado 0 (Raciones)', 'Grado 1,2,3 (Raciones)', 'Grado 4,5 (Raciones)',
             'Grado 6,7,8,9 (Raciones)', 'Grado 10,11 (Raciones)']
        ):
            insumos.extend(calcular_insumos(row['Complemento Alimenticio'], row['Ciclo de Menú (Semana)'], grupo, row[col_raciones]))

        resultados.append(insumos)

    # Generar PDFs con toda la información
    return generar_documentos(df)