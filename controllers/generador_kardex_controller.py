import os
import pandas as pd
import openpyxl
import win32com.client as win32
from tkinter import Tk, filedialog

# Rutas fijas según tu estructura
PLANTILLA_KARDEX = r"C:\Users\USUARIO\PAE_APP\v2\pae_web_adminlte_template\plantillas\kardex.xlsx"
DATOS_BASE = r"C:\Users\USUARIO\PAE_APP\v2\pae_web_adminlte_template\datos_base_completo_final_v2.xlsx"


def seleccionar_directorio_guardado():
    """Permite al usuario elegir dónde guardar el PDF final."""
    Tk().withdraw()
    carpeta = filedialog.askdirectory(title="Seleccione carpeta para guardar el Kardex")
    return carpeta


def calcular_insumos_kardex(complemento, ciclo, grupos_raciones):
    """
    Calcula insumos usando la hoja 'menus' del archivo base.
    grupos_raciones: dict con {'PREESCOLAR': x, 'PRIMARIA_BAJA': y, ...}
    """
    df_menus = pd.read_excel(DATOS_BASE, sheet_name="menus")

    resultados = []
    for grupo, raciones in grupos_raciones.items():
        filtro = (df_menus['componente'] == complemento) & \
                 (df_menus['menu'] == ciclo) & \
                 (df_menus['grupo'] == grupo)
        seleccion = df_menus.loc[filtro].copy()

        if not seleccion.empty:
            seleccion['cantidad_total'] = seleccion['cantidad_por_racion'] * raciones
            resultados.append(seleccion[['categoria', 'alimento', 'codigo',
                                         'cantidad_por_racion', 'unidad', 'cantidad_total']])

    if not resultados:
        return pd.DataFrame(columns=['categoria', 'alimento', 'codigo',
                                      'cantidad_por_racion', 'unidad', 'cantidad_total'])
    return pd.concat(resultados).groupby(['categoria', 'alimento', 'codigo', 'unidad'],
                                          as_index=False).sum(numeric_only=True)


def generar_kardex(df):
    """
    Genera un único PDF 'Kardex_PAE.pdf' consolidado, con un Kardex por institución.
    """
    carpeta_salida = seleccionar_directorio_guardado()
    if not carpeta_salida:
        return "Operación cancelada."

    archivos_temporales = []
    for idx, row in df.iterrows():
        # Cargar plantilla Kardex
        wb = openpyxl.load_workbook(PLANTILLA_KARDEX)
        ws = wb.active

        # Insertar datos generales
        ws['K2'] = pd.to_datetime(row['Fecha Inicial']).strftime("%d/%m/%Y")
        ws['I3'] = row['Municipio']
        ws['C4'] = row['Institucion Educativa (IED)']
        ws['J4'] = row['Sede']
        ws['C5'] = row['Complemento Alimenticio']
        ws['C6'] = row['Grado 0 (Raciones)']
        ws['F6'] = row['Grado 1,2,3 (Raciones)']
        ws['I6'] = row['Grado 4,5 (Raciones)']
        ws['L6'] = row['Grado 6,7,8,9 (Raciones)']
        ws['O6'] = row['Grado 10,11 (Raciones)']
        ws['F7'] = row['Cupos Totales']
        ws['J6'] = row['Ciclo de Menú (Semana)']

        # Calcular insumos dinámicos
        grupos = {
            'PREESCOLAR': row['Grado 0 (Raciones)'],
            'PRIMARIA_BAJA': row['Grado 1,2,3 (Raciones)'],
            'PRIMARIA_ALTA': row['Grado 4,5 (Raciones)'],
            'SECUNDARIA': row['Grado 6,7,8,9 (Raciones)'],
            'MYC': row['Grado 10,11 (Raciones)']
        }
        insumos = calcular_insumos_kardex(row['Complemento Alimenticio'],
                                          row['Ciclo de Menú (Semana)'],
                                          grupos)

        # Insertar alimentos dinámicamente desde fila 9
        fila_inicio = 9
        for _, alimento in insumos.iterrows():
            ws.cell(row=fila_inicio, column=2).value = alimento['alimento']  # Columna B
            ws.cell(row=fila_inicio, column=3).value = round(alimento['cantidad_total'], 2)  # Columna C
            ws.cell(row=fila_inicio, column=4).value = alimento['unidad']  # Columna D
            fila_inicio += 1

        # Guardar archivo temporal
        temp_file = os.path.join(carpeta_salida, f"Kardex_{idx+1}.xlsx")
        wb.save(temp_file)
        archivos_temporales.append(temp_file)

    # Combinar todos los Kardex en un único PDF
    pdf_salida = os.path.join(carpeta_salida, "Kardex_PAE.pdf")
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    merged_wb = excel.Workbooks.Add()

    for temp in archivos_temporales:
        wb_temp = excel.Workbooks.Open(os.path.abspath(temp))
        wb_temp.Sheets(1).Copy(Before=merged_wb.Sheets(1))
        wb_temp.Close(False)

    merged_wb.ExportAsFixedFormat(0, os.path.abspath(pdf_salida))
    merged_wb.Close(False)
    excel.Quit()

    # Limpiar temporales
    for temp in archivos_temporales:
        os.remove(temp)

    return f"Kardex generado correctamente en {pdf_salida}"